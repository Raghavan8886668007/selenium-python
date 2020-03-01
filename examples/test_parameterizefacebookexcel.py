from selenium.webdriver import Chrome
import openpyxl
path = "D://python-3.6.5//chromedriver_win32 (2)//chromedriver.exe"
driver = Chrome(executable_path=path)
driver.get("https://www.facebook.com/")
driver.maximize_window()

place = "D:/Raghavan/selenium/Demo.xlsx"
workbook = openpyxl.load_workbook(place)
sheet = workbook.active
rw = sheet.max_row
cl = sheet.max_column

for r in range(2,rw+1):
    for c in range(1,cl+1):
        row = r
        column = c
        #print(sheet.cell(r,c).value)

        driver.find_element_by_xpath("//input[@id='email']").send_keys(sheet.cell(row=2,column=1).value)
        driver.find_element_by_xpath("//input[@id='pass']").send_keys(sheet.cell(row=2,column=2).value)
        break
        print(sheet.cell(row=r,column = c).value)